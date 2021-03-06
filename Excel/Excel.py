from openpyxl import Workbook

wb = Workbook()


# ws1 = wb.active # 현재 활성화된 sheet 가져옴
# ws1.title = "sheet1" # 시트 이름 변경

ws = wb.create_sheet() # 새로운 기본 이름으로 생성
ws.title = "Test_Sheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "000000" #RGB 형태로 탭 색 변경

ws3 = wb.create_sheet("Test_Sheet3") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("Test_Sheet2", 2) # index에 Sheet 생성

new_ws = wb["Test_Sheet2"] # Dict 형태로 Sheet에 접근
target = wb.copy_worksheet(new_ws) # new_ws Sheet 복사
target.title = "COPY Sheet" # Sheet 이름 변경

print(wb.sheetnames) # 모든 Sheet 이름 확인

wb.save("sample.xlsx") # 저장
wb.close() # 닫기
